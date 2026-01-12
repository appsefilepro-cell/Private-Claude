"""
Data Ingestion Orchestrator
Handles ingestion from Gmail, Dropbox, OneDrive, SharePoint
Extracts customer contact details from emails, PDFs, and Excel sheets
"""

import os
import re
import csv
import json
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

# PDF processing
try:
    import fitz  # PyMuPDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logging.warning("PyMuPDF not available, PDF processing disabled")

# Excel processing
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl not available, Excel processing disabled")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger('IngestionOrchestrator')


class CustomerContact:
    """Customer contact data model"""

    def __init__(self):
        self.first_name = ""
        self.last_name = ""
        self.phone = ""
        self.email = ""
        self.mailing_address = ""
        self.tax_years = []
        self.filing_status = ""
        self.dependents = 0

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary"""
        return {
            "first_name": self.first_name,
            "last_name": self.last_name,
            "phone": self.phone,
            "email": self.email,
            "mailing_address": self.mailing_address,
            "tax_years": ", ".join(map(str, self.tax_years)),
            "filing_status": self.filing_status,
            "dependents": self.dependents
        }

    def to_csv_row(self) -> List[str]:
        """Convert to CSV row"""
        return [
            self.first_name,
            self.last_name,
            self.phone,
            self.email,
            self.mailing_address,
            ", ".join(map(str, self.tax_years)),
            self.filing_status,
            str(self.dependents)
        ]


class DataValidator:
    """Validates and normalizes data"""

    @staticmethod
    def validate_email(email: str) -> bool:
        """Validate email format"""
        pattern = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"
        return bool(re.match(pattern, email))

    @staticmethod
    def normalize_phone(phone: str) -> str:
        """Normalize phone to E.164 format"""
        # Remove all non-digit characters
        digits = re.sub(r'\D', '', phone)

        # Add country code if missing (assume US)
        if len(digits) == 10:
            digits = '1' + digits

        # Format to E.164
        if len(digits) == 11:
            return f"+{digits}"

        return phone  # Return original if can't normalize

    @staticmethod
    def extract_name(text: str) -> tuple:
        """Extract first and last name from text"""
        # Common patterns
        patterns = [
            r"Name:\s*(.+)",
            r"Full Name:\s*(.+)",
            r"Client Name:\s*(.+)",
            r"Taxpayer Name:\s*(.+)"
        ]

        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                name = match.group(1).strip()
                parts = name.split()
                if len(parts) >= 2:
                    return parts[0], parts[-1]

        return "", ""

    @staticmethod
    def extract_email(text: str) -> Optional[str]:
        """Extract email from text"""
        pattern = r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+"
        match = re.search(pattern, text)
        if match:
            email = match.group(0)
            if DataValidator.validate_email(email):
                return email
        return None

    @staticmethod
    def extract_phone(text: str) -> Optional[str]:
        """Extract phone number from text"""
        patterns = [
            r"Phone:\s*(.+)",
            r"Tel:\s*(.+)",
            r"Cell:\s*(.+)",
            r"\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}"
        ]

        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                phone = match.group(1) if 'Phone' in pattern or 'Tel' in pattern or 'Cell' in pattern else match.group(0)
                return DataValidator.normalize_phone(phone)

        return None


class PDFProcessor:
    """Process PDF files for customer data"""

    @staticmethod
    def extract_from_pdf(pdf_path: str) -> List[CustomerContact]:
        """Extract customer contacts from PDF"""
        if not PDF_AVAILABLE:
            logger.warning(f"PDF processing not available, skipping {pdf_path}")
            return []

        contacts = []

        try:
            doc = fitz.open(pdf_path)
            text = ""
            for page in doc:
                text += page.get_text()

            # Extract data
            contact = CustomerContact()

            # Name
            first, last = DataValidator.extract_name(text)
            contact.first_name = first
            contact.last_name = last

            # Email
            email = DataValidator.extract_email(text)
            if email:
                contact.email = email

            # Phone
            phone = DataValidator.extract_phone(text)
            if phone:
                contact.phone = phone

            # Address
            addr_match = re.search(r"Address:\s*(.+?)(?:\n|$)", text, re.IGNORECASE)
            if addr_match:
                contact.mailing_address = addr_match.group(1).strip()

            # Tax years
            tax_years = re.findall(r"Tax Year:\s*(\d{4})", text, re.IGNORECASE)
            if tax_years:
                contact.tax_years = [int(year) for year in tax_years]

            # Filing status
            status_match = re.search(r"Filing Status:\s*(.+?)(?:\n|$)", text, re.IGNORECASE)
            if status_match:
                contact.filing_status = status_match.group(1).strip()

            # Dependents
            dep_match = re.search(r"Dependents:\s*(\d+)", text, re.IGNORECASE)
            if dep_match:
                contact.dependents = int(dep_match.group(1))

            if contact.email or contact.phone:  # Only add if we have contact info
                contacts.append(contact)
                logger.info(f"Extracted contact from PDF: {contact.first_name} {contact.last_name}")

        except Exception as e:
            logger.error(f"Error processing PDF {pdf_path}: {e}")

        return contacts


class ExcelProcessor:
    """Process Excel files for customer data"""

    @staticmethod
    def extract_from_excel(excel_path: str) -> List[CustomerContact]:
        """Extract customer contacts from Excel"""
        if not EXCEL_AVAILABLE:
            logger.warning(f"Excel processing not available, skipping {excel_path}")
            return []

        contacts = []

        try:
            wb = load_workbook(excel_path)
            sheet = wb.active

            # Assume first row is headers
            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                contact = CustomerContact()

                # Map columns (flexible mapping based on header names)
                for idx, value in enumerate(row):
                    if idx >= len(headers) or not headers[idx]:
                        continue

                    header = str(headers[idx]).lower()

                    if 'first' in header and 'name' in header:
                        contact.first_name = str(value) if value else ""
                    elif 'last' in header and 'name' in header:
                        contact.last_name = str(value) if value else ""
                    elif 'email' in header:
                        if value and DataValidator.validate_email(str(value)):
                            contact.email = str(value)
                    elif 'phone' in header:
                        if value:
                            contact.phone = DataValidator.normalize_phone(str(value))
                    elif 'address' in header:
                        contact.mailing_address = str(value) if value else ""
                    elif 'tax' in header and 'year' in header:
                        if value:
                            contact.tax_years = [int(value)]
                    elif 'filing' in header and 'status' in header:
                        contact.filing_status = str(value) if value else ""
                    elif 'dependent' in header:
                        if value:
                            try:
                                contact.dependents = int(value)
                            except Exception:
                                contact.dependents = 0

                if contact.email or contact.phone:  # Only add if we have contact info
                    contacts.append(contact)
                    logger.info(f"Extracted contact from Excel: {contact.first_name} {contact.last_name}")

        except Exception as e:
            logger.error(f"Error processing Excel {excel_path}: {e}")

        return contacts


class IngestionOrchestrator:
    """Main orchestrator for data ingestion"""

    CSV_HEADERS = [
        "First Name", "Last Name", "Phone", "Email",
        "Mailing Address", "Tax Years", "Filing Status", "Dependents"
    ]

    def __init__(self, output_file: str = "customer_contact_list.csv"):
        self.output_file = output_file
        self.log_file = "logs/ingestion_log.json"
        self.contacts = []
        self.stats = {
            "total_processed": 0,
            "pdf_processed": 0,
            "excel_processed": 0,
            "contacts_extracted": 0,
            "duplicates_removed": 0,
            "errors": 0
        }

        # Ensure output CSV exists with headers
        self._init_csv()

        logger.info("Ingestion Orchestrator initialized")

    def _init_csv(self) -> None:
        """Initialize CSV file with headers if it doesn't exist"""
        if not os.path.exists(self.output_file):
            with open(self.output_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(self.CSV_HEADERS)
            logger.info(f"Created new CSV: {self.output_file}")

    def process_directory(self, directory: str) -> None:
        """Process all files in a directory"""
        logger.info(f"Processing directory: {directory}")

        if not os.path.exists(directory):
            logger.warning(f"Directory does not exist: {directory}")
            return

        for root, dirs, files in os.walk(directory):
            for file in files:
                file_path = os.path.join(root, file)
                self.process_file(file_path)

    def process_file(self, file_path: str) -> None:
        """Process a single file"""
        self.stats["total_processed"] += 1

        try:
            ext = Path(file_path).suffix.lower()

            if ext == '.pdf':
                contacts = PDFProcessor.extract_from_pdf(file_path)
                self.contacts.extend(contacts)
                self.stats["pdf_processed"] += 1
                self.stats["contacts_extracted"] += len(contacts)

            elif ext in ['.xlsx', '.xls']:
                contacts = ExcelProcessor.extract_from_excel(file_path)
                self.contacts.extend(contacts)
                self.stats["excel_processed"] += 1
                self.stats["contacts_extracted"] += len(contacts)

            else:
                logger.debug(f"Skipping unsupported file type: {file_path}")

        except Exception as e:
            logger.error(f"Error processing file {file_path}: {e}")
            self.stats["errors"] += 1

    def remove_duplicates(self) -> None:
        """Remove duplicate contacts based on email/phone"""
        unique_contacts = []
        seen = set()

        for contact in self.contacts:
            # Create unique key from email and phone
            key = f"{contact.email}|{contact.phone}"

            if key not in seen:
                seen.add(key)
                unique_contacts.append(contact)
            else:
                self.stats["duplicates_removed"] += 1

        self.contacts = unique_contacts
        logger.info(f"Removed {self.stats['duplicates_removed']} duplicates")

    def save_to_csv(self) -> None:
        """Save all contacts to CSV"""
        try:
            with open(self.output_file, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for contact in self.contacts:
                    writer.writerow(contact.to_csv_row())

            logger.info(f"Saved {len(self.contacts)} contacts to {self.output_file}")

        except Exception as e:
            logger.error(f"Error saving to CSV: {e}")

    def save_log(self) -> None:
        """Save ingestion log"""
        log_entry = {
            "timestamp": datetime.now().isoformat(),
            "stats": self.stats,
            "output_file": self.output_file
        }

        try:
            os.makedirs(os.path.dirname(self.log_file), exist_ok=True)

            # Append to log file
            logs = []
            if os.path.exists(self.log_file):
                with open(self.log_file, 'r') as f:
                    logs = json.load(f)

            logs.append(log_entry)

            with open(self.log_file, 'w') as f:
                json.dump(logs, f, indent=2)

            logger.info(f"Log saved to {self.log_file}")

        except Exception as e:
            logger.error(f"Error saving log: {e}")

    def run(self, directories: List[str]) -> Dict[str, Any]:
        """Run the full ingestion process"""
        logger.info("=== Starting Data Ingestion ===")

        # Process all directories
        for directory in directories:
            self.process_directory(directory)

        # Remove duplicates
        self.remove_duplicates()

        # Save results
        self.save_to_csv()
        self.save_log()

        logger.info("=== Ingestion Complete ===")
        logger.info(f"Stats: {json.dumps(self.stats, indent=2)}")

        return self.stats


def main():
    """Main entry point"""
    # Directories to process
    directories = [
        "data/gmail_attachments",
        "data/dropbox",
        "data/onedrive",
        "data/sharepoint",
        "data/local_files"
    ]

    orchestrator = IngestionOrchestrator()
    stats = orchestrator.run(directories)

    print("\n=== Ingestion Complete ===")
    print(json.dumps(stats, indent=2))


if __name__ == "__main__":
    main()
